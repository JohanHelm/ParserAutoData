from dataclasses import dataclass
from pathlib import Path
from openpyxl import load_workbook


@dataclass
class DataString:
    drom_brand: str
    drom_model: str
    dvorniki_model: str
    final_brand: str
    final_model: str
    extra_models: str


models_file = f"{Path(__file__).resolve().parent.parent}/TZ/good_models.xlsx"

workbook = load_workbook(filename=models_file, read_only=True, data_only=True)
sheet = workbook.active

rows_iterator = sheet.iter_rows(min_row=2,
                                max_row=2229,
                                min_col=1,
                                max_col=6,
                                values_only=True
                                )

converts_row_to_datastring = map(lambda value: DataString(*value), rows_iterator)
cut_empty_final_model = filter(lambda item: item.final_model, converts_row_to_datastring)
cut_extra_models = filter(lambda item: not item.extra_models, cut_empty_final_model)
cut_empty_dvorniki_model = filter(lambda item: item.dvorniki_model, cut_extra_models)
cut_equal_dvorniki_final_model = filter(lambda item: item.dvorniki_model != item.final_model, cut_empty_dvorniki_model)

result = {}
for data_string in cut_equal_dvorniki_final_model:
    models_dict = result.get(data_string.final_brand)
    if models_dict:
        models_dict[str(data_string.dvorniki_model)] = data_string.final_model
    else:
        models_dict = {data_string.final_brand: {str(data_string.dvorniki_model): data_string.final_model}}
        result.update(models_dict)
print(result)



